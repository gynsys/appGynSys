import openpyxl

wb_value = openpyxl.load_workbook("app diseño de mezclas.xlsx", data_only=True)
sheet = wb_value["DISEÑO IDEAL"]

print("--- ROWS 82-90 COLUMNS U to AE ---")
for r in range(82, 91):
    for c in range(21, 32): # Columns U to AE
        cell_ref = f"{openpyxl.utils.get_column_letter(c)}{r}"
        val = sheet.cell(row=r, column=c).value
        if val is not None:
            print(f"Cell {cell_ref}: {val}")
